import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import streamlit as st
import io

# -----------------------------
# Streamlit�̉�ʐݒ�
# -----------------------------
st.title("�������쐬")
date = st.text_input("�쐬�������������̓��t����͂��Ă�������(MM/DD)")
text_input = st.text_area("��������e�L�X�g����͂��Ă�������: ")

# �e�L�X�g���s�P�ʂŕ���
lines = text_input.splitlines()

# -----------------------------
# �e�픻��E���o�֐��̒�`
# -----------------------------

# �h���C�o�[���̊J�n�s���𔻒�
def is_driver_start_line(line):
    return line.count('��') >= 2

# �����𒊏o�i����2���� or �J�^�J�i3�����j
def extract_name(line):
    pattern = r'[��-?]{2}|[�@-��]{3}'
    match = re.findall(pattern, line)
    return match[0] if match else ""

# �����s�i���{�����j���ǂ���
def is_seikyu_line(line):
    return '��' in line and '����' in line

# �����z�ƍ�����𒊏o�i"123,456�~" �`���j
def is_seikyu_kousoku(line):
    pattern = r'(\d{1,3}(?:,\d{3})*)�~'
    match = re.findall(pattern, line)
    seikyu = int(match[0].replace(',', ''))
    kousoku = int(match[1].replace(',', ''))
    return seikyu, kousoku

# ���`���ɐ��`
def format_yen(value):
    return f"��{int(value):,}"

# -----------------------------
# �����f�[�^�̒��o���W�b�N
# -----------------------------
drivers_data = []
i = 0
while i < len(lines):
    line = lines[i]
    if is_driver_start_line(line):
        name = extract_name(line)
        j = i + 1
        while j < len(lines) and not is_driver_start_line(lines[j]):
            line = lines[j]
            if is_seikyu_line(line):
                seikyu, kousoku = is_seikyu_kousoku(line)
                drivers_data.append([name, seikyu, kousoku])
            j += 1
        # ���o�ςݕ������폜���ă��X�g��O�l�߂�
        del lines[i:j]
        i = 0
    else:
        i += 1

# -----------------------------
# Excel�t�@�C���̍쐬
# -----------------------------
wb = Workbook()
ws = wb.active
ws.title = "������"

# �^�C�g���s�̍쐬
ws.merge_cells("A1:J1")
ws["A1"] = f"������{date}"
ws["A1"].font = Font(size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A2"] = "���L�̒ʂ�Ɍ䐿���\���グ�܂�"

# �a��`���œ��t�o��
if "/" in date:
    month, day = date.split("/")
else:
    st.error("���t��MM/DD�̌`���œ��͂��Ă�������")
    st.stop()

if int(month) + 1 < 13:
    ws["I2"] = f"�ߘa7�N{int(month)+1}��1��"
else:
    ws["I2"] = f"�ߘa8�N{int(month)-11}��1��"

# �����z�Z���i���v�j
ws.merge_cells("A3:C3")
ws["A3"] = "�䐿���z"
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("D3:J3")
ws["D3"] = f"=SUM(J5:J{5+len(drivers_data)})"
ws["D3"].number_format = '"��"#,##0'
ws["D3"].font = Font(size=16, bold=True)
ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

# �w�b�_�[�s
ws.append([
    "���t", "�j��", "�Č���", "���e", "�S��", "�Ԏ�", "�P��", "�����", "������", "���v"
])

# -----------------------------
# �e�h���C�o�[�̃f�[�^��ǉ�
# -----------------------------
row_num = 5
for driver in drivers_data:
    name, seikyuu, kousoku = driver
    goukei = int(seikyuu) * 1.1 + int(kousoku)

    ws.append([
        f"2025/{date}",
        "=CHOOSE(WEEKDAY(A5, 1), \"��\", \"��\", \"��\", \"��\", \"��\", \"��\", \"�y\")",
        "���DC",
        "���ԑ݂��A��",
        name,
        "4t�①",
        int(seikyuu),
        f"=G{row_num}*0.1",
        int(kousoku),
        f"=G{row_num}+H{row_num}+I{row_num}"
    ])
    row_num += 1

# -----------------------------
# �g���iBorder�j�̐ݒ�
# -----------------------------
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

for row in ws.iter_rows(min_row=3, max_row=3+len(drivers_data)+1, min_col=1, max_col=10):
    for cell in row:
        cell.border = border

# -----------------------------
# Excel�̕ۑ����_�E�����[�h����
# -----------------------------
excel_buffer = io.BytesIO()
wb.save(excel_buffer)
excel_buffer.seek(0)

safe = date.replace('/', '_')  # �t�@�C�����p�̃X���b�V���u��

st.download_button(
    label="�������쐬",
    data=excel_buffer,
    file_name=f"{safe}������.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
